# -*- coding: utf-8 -*-
import openpyxl
from datetime import datetime, timedelta

EXCEL_PATH = r'd:\Odoo\bmad-odoo\_bmad-output\Tài liệu\Doanh thu\Báo cáo Doanh thu\Báo cáo Doanh thu\Đối_chiếu_doanh_thu_2026 v1.xlsx'
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['Đối chiếu']

BASE = datetime(1899, 12, 30)

def serial_to_date(serial):
    if serial and isinstance(serial, (int, float)):
        return BASE + timedelta(days=int(serial))
    return None

# Header
print("=== Header ===")
header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print(header_row)

# Find all unique dates
dates = set()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=6):
    for cell in row:
        d = serial_to_date(cell.value)
        if d:
            dates.add(d)

print(f"\n=== Unique dates ({len(dates)}) ===")
for d in sorted(dates):
    fmt = d.strftime("%Y-%m-%d")
    print(f"  {fmt} (day={d.day})")

# Now filter rows where ngay_odoo day is 11-20
print("\n=== Rows with ngay_odoo day 11-20 (showing sai_lech != 'Khớp' and != 'Lệch ngày') ===")
count = 0
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    vals = [cell.value for cell in row]
    # cols: 0=ma_don, 1=so_tai_khoan, 2=noi_dung_odoo, 3=noi_dung_ke_toan, 
    #        4=ngay_odoo, 5=ngay_ke_toan, 6=tien_odoo, 7=tien_ke_toan, 8=sai_lech, 9=col_J
    ngay_odoo = serial_to_date(vals[4])
    if ngay_odoo and 11 <= ngay_odoo.day <= 20:
        sai_lech = vals[8] or ''
        # Only orange (Lệch số tiền) and pink (Không khớp)
        if 'Khớp' == sai_lech.strip() or sai_lech.startswith('Lệch ngày'):
            continue
        ngay_kt = serial_to_date(vals[5])
        ngay_odoo_str = ngay_odoo.strftime("%d/%m/%Y") if ngay_odoo else ''
        ngay_kt_str = ngay_kt.strftime("%d/%m/%Y") if ngay_kt else ''
        col_j = vals[9] if len(vals) > 9 else ''
        
        print(f"Row {row_idx}: {vals[0]} | TK={vals[1]} | Odoo={vals[2]} | KT={vals[3]} | "
              f"ngay_odoo={ngay_odoo_str} | ngay_kt={ngay_kt_str} | "
              f"tien_odoo={vals[6]} | tien_kt={vals[7]} | sai_lech={sai_lech} | col_J={col_j}")
        count += 1

print(f"\nTotal matching rows: {count}")
