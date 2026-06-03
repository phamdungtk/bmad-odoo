# -*- coding: utf-8 -*-
import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict

EXCEL_PATH = r'd:\Odoo\bmad-odoo\_bmad-output\Tài liệu\Doanh thu\Báo cáo Doanh thu\Báo cáo Doanh thu\Đối_chiếu_doanh_thu_2026 v1.xlsx'
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['Đối chiếu']

BASE = datetime(1899, 12, 30)

def serial_to_date(serial):
    if serial and isinstance(serial, (int, float)):
        return BASE + timedelta(days=int(serial))
    return None

# Check what data exists around the user's screenshots (rows 2267-2296)
print("=== Rows 2265-2300 (around screenshots) ===")
for row_idx, row in enumerate(ws.iter_rows(min_row=2265, max_row=2300), start=2265):
    vals = [cell.value for cell in row]
    ngay_odoo = serial_to_date(vals[4])
    ngay_kt = serial_to_date(vals[5])
    no = ngay_odoo.strftime("%d/%m/%Y") if ngay_odoo else ''
    nk = ngay_kt.strftime("%d/%m/%Y") if ngay_kt else ''
    sai = vals[8] or ''
    j = vals[9] or ''
    print(f"Row {row_idx}: {vals[0]} | TK={vals[1]} | ngay_o={no} | ngay_kt={nk} | "
          f"t_o={vals[6]} | t_kt={vals[7]} | sai={sai[:50]} | J={j}")

# Now let's look at the existing col J patterns for rows that ALREADY have values
print("\n\n=== All rows with existing col_J values ===")
j_patterns = defaultdict(int)
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    vals = [cell.value for cell in row]
    col_j = vals[9] if len(vals) > 9 else None
    if col_j:
        j_patterns[str(col_j)[:60]] += 1
        
print("Col J patterns found:")
for pattern, count in sorted(j_patterns.items(), key=lambda x: -x[1]):
    print(f"  [{count}x] {pattern}")

# Show some examples of rows with col_J filled in
print("\n=== Sample rows WITH col_J (first 30) ===")
shown = 0
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    vals = [cell.value for cell in row]
    col_j = vals[9] if len(vals) > 9 else None
    if col_j:
        ngay_odoo = serial_to_date(vals[4])
        ngay_kt = serial_to_date(vals[5])
        no = ngay_odoo.strftime("%d/%m/%Y") if ngay_odoo else ''
        nk = ngay_kt.strftime("%d/%m/%Y") if ngay_kt else ''
        sai = vals[8] or ''
        print(f"Row {row_idx}: {vals[0]} | TK={vals[1]} | ngay_o={no} | ngay_kt={nk} | "
              f"t_o={vals[6]} | t_kt={vals[7]} | sai={sai[:60]} | J={col_j}")
        shown += 1
        if shown >= 30:
            break
