# -*- coding: utf-8 -*-
"""
Analyze rows with ngay_odoo from day 11-20, only orange/pink rows.
Group by ma_don to understand KT grouping patterns.
"""
import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict

EXCEL_PATH = r'd:\Odoo\bmad-odoo\_bmad-output\Tài liệu\Doanh thu\Báo cáo Doanh thu\Báo cáo Doanh thu\Đối_chiếu_doanh_thu_2026 v1.xlsx'
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['Đối chiếu']

BASE = datetime(1899, 12, 30)

def serial_to_date(serial):
    if serial and isinstance(serial, (int, float)):
        return BASE + timedelta(days=int(serial))
    return None

# Collect ALL rows for context, grouped by ma_don
all_rows = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    vals = [cell.value for cell in row]
    all_rows.append((row_idx, vals))

# Group by ma_don
by_don = defaultdict(list)
for row_idx, vals in all_rows:
    ma_don = vals[0]
    if ma_don:
        by_don[ma_don].append((row_idx, vals))

# Now focus on rows with ngay_odoo day 11-20, only Lệch/Không khớp
print("=" * 120)
print("PHÂN TÍCH DÒNG CAM/HỒNG VỚI ngay_odoo NGÀY 11-20 (MỌI THÁNG)")
print("=" * 120)

target_rows = []
for row_idx, vals in all_rows:
    ngay_odoo = serial_to_date(vals[4])
    if ngay_odoo and 11 <= ngay_odoo.day <= 20:
        sai_lech = str(vals[8] or '')
        # Only orange (Lệch số tiền) and pink (Không khớp)
        if sai_lech == 'Khớp' or sai_lech.startswith('Lệch ngày:'):
            continue
        col_j = vals[9] if len(vals) > 9 else None
        target_rows.append((row_idx, vals, ngay_odoo))

# Group target rows by ma_don to understand patterns
print(f"\nTotal target rows: {len(target_rows)}")

# Group by ma_don
target_by_don = defaultdict(list)
for row_idx, vals, ngay_odoo in target_rows:
    target_by_don[vals[0]].append((row_idx, vals, ngay_odoo))

print(f"Unique ma_don: {len(target_by_don)}")

# For each ma_don with target rows, show ALL rows of that don for context
print("\n" + "=" * 120)
print("CHI TIẾT THEO ĐƠN (hiện tất cả dòng của đơn, đánh dấu ★ cho dòng target)")
print("=" * 120)

for ma_don in sorted(target_by_don.keys()):
    target_row_ids = {r[0] for r in target_by_don[ma_don]}
    all_don_rows = by_don[ma_don]
    
    print(f"\n{'─' * 100}")
    print(f"ĐƠN: {ma_don} ({len(all_don_rows)} dòng tổng, {len(target_row_ids)} dòng target)")
    print(f"{'─' * 100}")
    
    for row_idx, vals in all_don_rows:
        ngay_odoo = serial_to_date(vals[4])
        ngay_kt = serial_to_date(vals[5])
        no = ngay_odoo.strftime("%d/%m/%Y") if ngay_odoo else ''
        nk = ngay_kt.strftime("%d/%m/%Y") if ngay_kt else ''
        sai = str(vals[8] or '')[:70]
        col_j = vals[9] if len(vals) > 9 else None
        marker = "★" if row_idx in target_row_ids else " "
        
        print(f"  {marker} Row {row_idx}: TK={vals[1]:10s} | o={no:10s} | kt={nk:10s} | "
              f"t_o={str(vals[6]):>15s} | t_kt={str(vals[7]):>15s} | {sai}")
        if col_j:
            print(f"       → col_J: {col_j}")
