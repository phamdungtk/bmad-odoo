# -*- coding: utf-8 -*-
"""
Full analysis: rows with ngay_odoo day 11-20, only cam/hồng.
Output ALL groups to a file for review.
"""
import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict
import sys

EXCEL_PATH = r'd:\Odoo\bmad-odoo\_bmad-output\Tài liệu\Doanh thu\Báo cáo Doanh thu\Báo cáo Doanh thu\Đối_chiếu_doanh_thu_2026 v1.xlsx'
OUT_PATH = r'd:\Odoo\bmad-odoo\_bmad-output\Tài liệu\Doanh thu\Báo cáo Doanh thu\Báo cáo Doanh thu\analysis_output.txt'

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['Đối chiếu']

BASE = datetime(1899, 12, 30)

def serial_to_date(serial):
    if serial and isinstance(serial, (int, float)):
        return BASE + timedelta(days=int(serial))
    return None

def fmt_money(v):
    if v is None:
        return ''
    return f"{v:>15,.0f}"

# Collect ALL rows
all_rows = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    vals = [cell.value for cell in row]
    all_rows.append((row_idx, vals))

# Group by ma_don
by_don = defaultdict(list)
for row_idx, vals in all_rows:
    ma_don = vals[0]
    if ma_don:
        by_don[ma_don].append((row_idx, vals))

# Find target rows (ngay_odoo day 11-20, only cam/hồng)
target_rows = []
for row_idx, vals in all_rows:
    ngay_odoo = serial_to_date(vals[4])
    if ngay_odoo and 11 <= ngay_odoo.day <= 20:
        sai_lech = str(vals[8] or '')
        if sai_lech == 'Khớp' or sai_lech.startswith('Lệch ngày:'):
            continue
        target_rows.append((row_idx, vals, ngay_odoo))

# Group target rows by ma_don
target_by_don = defaultdict(list)
for row_idx, vals, ngay_odoo in target_rows:
    target_by_don[vals[0]].append((row_idx, vals, ngay_odoo))

# Write full output
with open(OUT_PATH, 'w', encoding='utf-8') as f:
    f.write("=" * 140 + "\n")
    f.write("PHÂN TÍCH DÒNG CAM/HỒNG VỚI ngay_odoo NGÀY 11-20\n")
    f.write(f"Tổng dòng target: {len(target_rows)}, Tổng đơn: {len(target_by_don)}\n")
    f.write("=" * 140 + "\n")
    
    # Group by month for clarity
    by_month = defaultdict(list)
    for ma_don in sorted(target_by_don.keys()):
        # Get earliest month from target rows
        earliest = min(r[2] for r in target_by_don[ma_don])
        month_key = earliest.strftime("%Y-%m")
        by_month[month_key].append(ma_don)
    
    for month_key in sorted(by_month.keys(), reverse=True):
        f.write(f"\n{'#' * 140}\n")
        f.write(f"# THÁNG {month_key} ({len(by_month[month_key])} đơn)\n")
        f.write(f"{'#' * 140}\n")
        
        for ma_don in sorted(by_month[month_key]):
            target_row_ids = {r[0] for r in target_by_don[ma_don]}
            all_don_rows = by_don[ma_don]
            
            f.write(f"\n{'─' * 130}\n")
            f.write(f"ĐƠN: {ma_don} ({len(all_don_rows)} dòng, {len(target_row_ids)} target)\n")
            f.write(f"{'─' * 130}\n")
            
            # Check if any row already has col_J
            has_col_j = any(v[9] for _, v in all_don_rows if len(v) > 9 and v[9])
            if has_col_j:
                f.write("  ℹ️  Đơn đã có col_J ghi nhận trước đó\n")
            
            for row_idx, vals in all_don_rows:
                ngay_odoo = serial_to_date(vals[4])
                ngay_kt = serial_to_date(vals[5])
                no = ngay_odoo.strftime("%d/%m/%Y") if ngay_odoo else ''
                nk = ngay_kt.strftime("%d/%m/%Y") if ngay_kt else ''
                sai = str(vals[8] or '')
                col_j = vals[9] if len(vals) > 9 else None
                marker = "★" if row_idx in target_row_ids else " "
                
                f.write(f"  {marker} Row {row_idx:4d}: TK={str(vals[1]):10s} | "
                        f"o={no:10s} | kt={nk:10s} | "
                        f"t_o={fmt_money(vals[6])} | t_kt={fmt_money(vals[7])} | "
                        f"{sai}\n")
                if col_j:
                    f.write(f"       → J: {col_j}\n")
            
            # Attempt to detect pattern
            # Pattern 1: Same don, KT groups multiple Odoo lines into one
            # Pattern 2: KT recorded in different month
            target_list = target_by_don[ma_don]
            lech_rows = [(r, v) for r, v, _ in target_list if 'Lệch số tiền' in str(v[8] or '')]
            khong_khop_rows = [(r, v) for r, v, _ in target_list if 'Không khớp' in str(v[8] or '')]
            
            # Detect if KT date is in a different month
            for row_idx, vals, ngay_odoo in target_list:
                ngay_kt = serial_to_date(vals[5])
                if ngay_kt and ngay_odoo.month != ngay_kt.month:
                    kt_str = ngay_kt.strftime("%d/%m")
                    f.write(f"  🔎 PHÁT HIỆN: Row {row_idx} - KT ghi nhận ở tháng khác ({kt_str} vs Odoo {ngay_odoo.strftime('%d/%m')})\n")
                    f.write(f"     → GỢI Ý col_J: KT gom lại ghi vào {kt_str}, HT ghi nhận sau\n")
            
            # Detect if it's a KT grouping pattern (one row has lệch, companion row has matching amount)
            for r1, v1 in lech_rows:
                lech_amount = 0
                sai_str = str(v1[8] or '')
                # Extract lệch amount from sai_lech string
                import re
                match = re.search(r'Lệch số tiền: ([+-]?[\d,]+)', sai_str)
                if match:
                    lech_amount = int(match.group(1).replace(',', '').replace('+', ''))
                
                for r2, v2 in khong_khop_rows:
                    odoo2 = v2[6] or 0
                    if abs(abs(lech_amount) - abs(odoo2)) < 10:  # within rounding
                        f.write(f"  🔎 PHÁT HIỆN: Row {r1} (lệch {lech_amount:,.0f}) + Row {r2} (Odoo {odoo2:,.0f}) = KT GOM\n")
                        
                        # Check if cross-month
                        ngay_odoo_1 = serial_to_date(v1[4])
                        ngay_kt_1 = serial_to_date(v1[5])
                        if ngay_kt_1 and ngay_odoo_1 and ngay_odoo_1.month != ngay_kt_1.month:
                            kt_str = ngay_kt_1.strftime("%d/%m")
                            f.write(f"     → GỢI Ý col_J: KT gom lại ghi vào {kt_str}, HT ghi nhận sau\n")
                        else:
                            f.write(f"     → GỢI Ý col_J: KT gom vào ghi nhận\n")

print(f"Output written to: {OUT_PATH}")
print(f"Total target rows: {len(target_rows)}")
print(f"Total unique ma_don: {len(target_by_don)}")

# Also print summary stats
sai_types = defaultdict(int)
for _, vals, _ in target_rows:
    sai = str(vals[8] or '')
    if 'Lệch số tiền' in sai and 'Lệch ngày' in sai:
        sai_types['Lệch số tiền + Lệch ngày'] += 1
    elif 'Lệch số tiền' in sai:
        sai_types['Lệch số tiền'] += 1
    elif 'đơn không có trên kế toán' in sai:
        sai_types['Không khớp: đơn không có trên KT'] += 1
    elif 'không có dòng nội dung' in sai:
        sai_types['Không khớp: KT có đơn, thiếu nội dung'] += 1
    elif 'bút toán kế toán dư' in sai:
        sai_types['Không khớp: bút toán KT dư'] += 1
    else:
        sai_types[sai[:40]] += 1

print("\n=== Phân loại sai lệch ===")
for sai_type, count in sorted(sai_types.items(), key=lambda x: -x[1]):
    print(f"  [{count:3d}] {sai_type}")
