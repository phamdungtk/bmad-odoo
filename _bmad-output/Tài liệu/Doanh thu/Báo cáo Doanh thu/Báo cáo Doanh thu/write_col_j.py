# -*- coding: utf-8 -*-
"""
Script ghi cột J vào file Excel theo quy tắc đối chiếu doanh thu.
Chỉ xử lý dòng cam/hồng (Lệch số tiền / Không khớp) với ngay_odoo ngày 11-20.

QUY TẮC CỘT J:
1. "KT gom vào ghi nhận" — Khi phát hiện pattern KT gom:
   - Row A có "Lệch số tiền: -X" 
   - Row B cùng đơn có "Không khớp" với tien_odoo ≈ X
   → Cả 2 row ghi "KT gom vào ghi nhận"

2. "KT gom lại ghi vào DD/MM, HT ghi nhận sau" — Khi:
   - Row A có Lệch số tiền VÀ lệch ngày khác tháng (ngay_kt khác tháng ngay_odoo)
   - Row B cùng đơn "Không khớp"
   → Ghi "KT gom lại ghi vào DD/MM, HT ghi nhận sau"

3. "Không có trên Misa" — Khi đơn hoàn toàn "Không khớp: đơn không có trên kế toán"

4. "Lệch kỳ" — Khi chỉ lệch số tiền nhỏ (< 1000đ) do làm tròn

5. "Khớp" — Khi dòng target thuộc đơn mà tổng Odoo ≈ tổng KT (KT ghi đầy đủ ở bút toán 511)

6. "KT gom lại ghi vào DD/MM" — Khi KT date khác tháng, không kèm "HT ghi nhận sau"
"""
import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict
import re
import copy

EXCEL_PATH = r'd:\Odoo\bmad-odoo\_bmad-output\Tài liệu\Doanh thu\Báo cáo Doanh thu\Báo cáo Doanh thu\Đối_chiếu_doanh_thu_2026 v1.xlsx'
OUTPUT_PATH = r'd:\Odoo\bmad-odoo\_bmad-output\Tài liệu\Doanh thu\Báo cáo Doanh thu\Báo cáo Doanh thu\Đối_chiếu_doanh_thu_2026 v4.xlsx'

BASE = datetime(1899, 12, 30)

def serial_to_date(serial):
    if serial is None:
        return None
    if isinstance(serial, datetime):
        return serial
    if isinstance(serial, (int, float)):
        return BASE + timedelta(days=int(serial))
    return None

def extract_lech_amount(sai_str):
    """Extract lệch số tiền from sai_lech string."""
    match = re.search(r'Lệch số tiền: ([+-]?[\d,]+)', sai_str)
    if match:
        return int(match.group(1).replace(',', '').replace('+', ''))
    return None

# Load workbook WITH formatting (not data_only)
print("Loading workbook...")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['Đối chiếu']

# Also load data_only version for computed values
wb_data = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws_data = wb_data['Đối chiếu']

# Read all data
print("Reading data...")
all_rows = []
for row_idx, row in enumerate(ws_data.iter_rows(min_row=2, max_row=ws_data.max_row), start=2):
    vals = [cell.value for cell in row]
    all_rows.append((row_idx, vals))

# Group by ma_don (exact)
by_don = defaultdict(list)
for row_idx, vals in all_rows:
    ma_don = vals[0]
    if ma_don:
        by_don[ma_don].append((row_idx, vals))

# Build family groups: S14851, S14851_1, S14851_2 → all under "S14851"
def get_base_don(ma_don):
    """Extract base don name, stripping _1, _2, _3 suffixes."""
    match = re.match(r'^(S\d+)_\d+$', str(ma_don))
    return match.group(1) if match else str(ma_don)

by_don_family = defaultdict(list)
for row_idx, vals in all_rows:
    ma_don = vals[0]
    if ma_don:
        base = get_base_don(ma_don)
        by_don_family[base].append((row_idx, vals))

# Find target rows (ngay_odoo day 11-20, only Lệch/Không khớp)
target_rows = []
for row_idx, vals in all_rows:
    ngay_odoo = serial_to_date(vals[4])
    if ngay_odoo and 11 <= ngay_odoo.day <= 20:
        sai_lech = str(vals[8] or '')
        if sai_lech == 'Khớp' or sai_lech.startswith('Lệch ngày:'):
            continue
        target_rows.append((row_idx, vals, ngay_odoo))

# Group target rows by ma_don
target_by_don = defaultdict(list)
for row_idx, vals, ngay_odoo in target_rows:
    target_by_don[vals[0]].append((row_idx, vals, ngay_odoo))

print(f"Target rows: {len(target_rows)}, Unique đơn: {len(target_by_don)}")

# Process and assign col_J values
col_j_updates = {}  # row_idx -> col_J text
processed_count = 0

for ma_don, target_list in target_by_don.items():
    target_row_ids = {r[0] for r in target_list}
    base = get_base_don(ma_don)
    # Use family group (includes _1, _2 variants) for matching
    all_don_rows = by_don_family[base]
    
    # Separate target rows by type
    lech_rows = []  # Rows with "Lệch số tiền"
    khong_khop_rows = []  # Rows with "Không khớp"
    
    for row_idx, vals, ngay_odoo in target_list:
        sai = str(vals[8] or '')
        if 'Lệch số tiền' in sai:
            lech_rows.append((row_idx, vals, ngay_odoo))
        elif 'Không khớp' in sai:
            khong_khop_rows.append((row_idx, vals, ngay_odoo))
    
    # Skip rows that already have col_J
    already_has_j = set()
    for row_idx, vals, _ in target_list:
        existing_j = vals[9] if len(vals) > 9 else None
        if existing_j:
            already_has_j.add(row_idx)
    
    # ─── RULE 1: KT GOM pattern ───
    # One row has "Lệch số tiền: -X", companion row has "Không khớp" with tien_odoo ≈ |X|
    matched_lech = set()
    matched_khop = set()
    
    for r1, v1, d1 in lech_rows:
        lech_amount = extract_lech_amount(str(v1[8] or ''))
        if lech_amount is None:
            continue
        
        for r2, v2, d2 in khong_khop_rows:
            if r2 in matched_khop:
                continue
            odoo_amount = v2[6] or 0
            if abs(abs(lech_amount) - abs(odoo_amount)) < 100:  # within rounding tolerance
                # Found a KT GOM pair!
                ngay_kt = serial_to_date(v1[5])
                ngay_odoo = d1
                
                if ngay_kt and ngay_odoo and ngay_odoo.month != ngay_kt.month:
                    # Cross-month: "KT gom lại ghi vào DD/MM, HT ghi nhận sau"
                    kt_str = ngay_kt.strftime("%-d/%m").replace("%-d", str(ngay_kt.day))
                    label = f"KT gom lại ghi vào {ngay_kt.day}/{ngay_kt.month}, HT ghi nhận sau"
                else:
                    # Same month: "KT gom vào ghi nhận"
                    label = "KT gom vào ghi nhận"
                
                if r1 not in already_has_j:
                    col_j_updates[r1] = label
                if r2 not in already_has_j:
                    col_j_updates[r2] = label
                
                matched_lech.add(r1)
                matched_khop.add(r2)
                break
    
    # ─── RULE 2: Lệch số tiền nhỏ (làm tròn) ───
    for r1, v1, d1 in lech_rows:
        if r1 in matched_lech or r1 in already_has_j:
            continue
        lech_amount = extract_lech_amount(str(v1[8] or ''))
        if lech_amount is not None and abs(lech_amount) < 1000:
            col_j_updates[r1] = "Lệch làm tròn"
    
    # ─── RULE 3: "Không khớp: đơn không có trên kế toán" ───
    all_don_ko_co = all(
        'đơn không có trên kế toán' in str(v[8] or '')
        for _, v, _ in target_list
        if 'Không khớp' in str(v[8] or '')
    )
    
    for r2, v2, d2 in khong_khop_rows:
        if r2 in matched_khop or r2 in already_has_j:
            continue
        sai = str(v2[8] or '')
        
        if 'đơn không có trên kế toán' in sai:
            col_j_updates[r2] = "Không có trên Misa"
        elif 'không có dòng nội dung' in sai:
            odoo_amount = v2[6] or 0
            
            # ─── CHECK: suffix entries (_1, _2) have matching KT for this amount? ───
            suffix_match_ma = None
            for _, v_other in all_don_rows:
                other_ma = str(v_other[0] or '')
                other_sai = str(v_other[8] or '')
                other_kt_amount = v_other[7] or 0
                # Look in _1, _2 entries for matching KT amount
                if (other_ma != ma_don 
                    and get_base_don(other_ma) == base
                    and abs(other_kt_amount - odoo_amount) < 100 
                    and odoo_amount > 0):
                    suffix_match_ma = other_ma
                    break
            
            if odoo_amount == 0:
                # Odoo = 0, KT doesn't have it → basically OK
                col_j_updates[r2] = "Odoo = 0, bỏ qua"
            elif suffix_match_ma:
                # Found exact matching amount in _1/_2 suffix entry
                kt_suffix_date = None
                for _, v_s in all_don_rows:
                    if str(v_s[0] or '') == suffix_match_ma:
                        kt_suffix_date = serial_to_date(v_s[5])
                        if kt_suffix_date:
                            break
                if kt_suffix_date:
                    col_j_updates[r2] = f"KT ghi nhận tại {suffix_match_ma} ngày {kt_suffix_date.day}/{kt_suffix_date.month}"
                else:
                    col_j_updates[r2] = f"KT ghi nhận tại {suffix_match_ma}"
            else:
                # Check if there's a matching 511 "bút toán kế toán dư" row
                has_bt_du = any(
                    'bút toán kế toán dư' in str(v[8] or '')
                    for _, v in all_don_rows
                )
                
                if has_bt_du:
                    # KT has the amount in a 511 aggregate entry
                    # Find the 511 entry's date and amount
                    kt_511_date = None
                    kt_511_total = 0
                    for _, v_other in all_don_rows:
                        if 'bút toán kế toán dư' in str(v_other[8] or ''):
                            if kt_511_date is None:
                                kt_511_date = serial_to_date(v_other[5])
                            kt_511_total += (v_other[7] or 0)
                    
                    # Calculate total Odoo for "không có dòng nội dung" target rows of this don
                    odoo_target_total = sum(
                        (v[6] or 0)
                        for _, v, _ in target_list
                        if 'không có dòng nội dung' in str(v[8] or '') and (v[6] or 0) > 0
                    )
                    
                    # Check if amounts match
                    amounts_match = abs(odoo_target_total - kt_511_total) < 100  # tolerance
                    lech_suffix = "" if amounts_match else ", lệch số tiền"
                    
                    if kt_511_date:
                        col_j_updates[r2] = f"KT gom lại ghi vào {kt_511_date.day}/{kt_511_date.month}{lech_suffix}"
                    else:
                        col_j_updates[r2] = f"KT gom lại{lech_suffix}"
                else:
                    # No matching KT entry at all
                    # Check if Lệch rows exist for same don with cross-month
                    cross_month_date = None
                    for _, v_other in all_don_rows:
                        sai_other = str(v_other[8] or '')
                        if 'Lệch số tiền' in sai_other:
                            ngay_kt_other = serial_to_date(v_other[5])
                            ngay_odoo_other = serial_to_date(v_other[4])
                            if ngay_kt_other and ngay_odoo_other and ngay_odoo_other.month != ngay_kt_other.month:
                                cross_month_date = ngay_kt_other
                                break
                    
                    if cross_month_date:
                        col_j_updates[r2] = f"KT gom lại ghi vào {cross_month_date.day}/{cross_month_date.month}, HT ghi nhận sau"
                    else:
                        col_j_updates[r2] = "Khớp"
    
    # ─── RULE 4: Unmatched lệch rows with cross-month ───
    for r1, v1, d1 in lech_rows:
        if r1 in matched_lech or r1 in already_has_j or r1 in col_j_updates:
            continue
        ngay_kt = serial_to_date(v1[5])
        if ngay_kt and d1 and d1.month != ngay_kt.month:
            col_j_updates[r1] = f"KT gom lại ghi vào {ngay_kt.day}/{ngay_kt.month}, HT ghi nhận sau"

# Write col_J values to the formatted workbook
print(f"\nWriting {len(col_j_updates)} col_J values...")

for row_idx, col_j_text in sorted(col_j_updates.items()):
    cell = ws.cell(row=row_idx, column=10)
    cell.value = col_j_text
    processed_count += 1

# Save
print(f"Saving to {OUTPUT_PATH}...")
wb.save(OUTPUT_PATH)
print(f"Done! Processed {processed_count} cells.")

# Print summary
labels = defaultdict(int)
for text in col_j_updates.values():
    # Normalize for counting
    if text.startswith("KT gom lại ghi vào") and "HT ghi nhận sau" in text:
        labels["KT gom lại ghi vào DD/MM, HT ghi nhận sau"] += 1
    elif text.startswith("KT gom lại ghi vào") and "lệch số tiền" in text:
        labels["KT gom lại ghi vào DD/MM, lệch số tiền"] += 1
    elif text.startswith("KT gom lại ghi vào"):
        labels["KT gom lại ghi vào DD/MM"] += 1
    elif text.startswith("KT ghi nhận tại"):
        labels["KT ghi nhận tại _N"] += 1
    elif text == "KT gom vào ghi nhận":
        labels["KT gom vào ghi nhận"] += 1
    elif text == "Không có trên Misa":
        labels["Không có trên Misa"] += 1
    elif text == "Lệch làm tròn":
        labels["Lệch làm tròn"] += 1
    elif text == "Odoo = 0, bỏ qua":
        labels["Odoo = 0, bỏ qua"] += 1
    elif text == "Khớp":
        labels["Khớp"] += 1
    else:
        labels[text] += 1

print("\n=== Tổng hợp col_J đã ghi ===")
for label, count in sorted(labels.items(), key=lambda x: -x[1]):
    print(f"  [{count:3d}] {label}")

# Print all updates for review
print("\n=== Chi tiết tất cả updates ===")
for row_idx in sorted(col_j_updates.keys()):
    vals = None
    for r, v in all_rows:
        if r == row_idx:
            vals = v
            break
    if vals:
        ngay_odoo = serial_to_date(vals[4])
        no = ngay_odoo.strftime("%d/%m/%Y") if ngay_odoo else ''
        sai = str(vals[8] or '')[:55]
        print(f"  Row {row_idx:4d}: {vals[0]} | TK={vals[1]:10s} | ngay_o={no} | "
              f"t_o={vals[6] or 0:>15,.0f} | {sai}")
        print(f"       → J: {col_j_updates[row_idx]}")
